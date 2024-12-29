from django.test import TestCase
from .services.report_generator import calculate_tax
import os
import pandas as pd
from openpyxl import load_workbook
from .services.report_generator import generate_report
import tempfile
from .services.file_parser import parse_file
from openpyxl import Workbook


class CalculateTaxTests(TestCase):
    def test_zero_tax_base(self):
        """Проверка нулевой налоговой базы"""
        self.assertEqual(calculate_tax(0), 0)

    def test_small_positive_tax_base(self):
        """Проверка малой положительной налоговой базы"""
        self.assertEqual(calculate_tax(1000000), 130000)

    def test_upper_limit_of_13_percent_tax(self):
        """Проверка верхней границы для ставки 13%"""
        self.assertEqual(calculate_tax(5000000), 650000)

    def test_lower_limit_of_15_percent_tax(self):
        """Проверка нижней границы для ставки 15%"""
        self.assertEqual(calculate_tax(5000001), 650000 + 0.15)

    def test_large_tax_base(self):
        """Проверка большой налоговой базы"""
        self.assertEqual(calculate_tax(10000000), 650000 + (5000000 * 0.15))


class GenerateReportTests(TestCase):
    def setUp(self):
        self.valid_data = pd.DataFrame(
            {
                "Филиал": ["Филиал 1", "Филиал 2"],
                "Сотрудник": ["Сотрудник 1", "Сотрудник 2"],
                "Налоговая база": [5000000, 6000000],
                "Исчислено всего": [650000, 800000],
            }
        )

    def test_report_creation(self):
        """Проверка, что отчет создается"""
        report_path = generate_report(self.valid_data)
        self.assertTrue(os.path.exists(report_path))
        os.remove(report_path)

    def test_report_structure(self):
        """Проверка структуры отчета"""
        report_path = generate_report(self.valid_data)
        workbook = load_workbook(report_path)
        sheet = workbook.active
        if sheet:
            expected_headers = [
                ("A1", "Филиал"),
                ("B1", "Сотрудник"),
                ("C1", "Налоговая база"),
                ("D1", "Налог"),
                ("D2", "Исчислено всего"),
                ("E2", "Исчислено всего по формуле"),
                ("F1", "Отклонения"),
            ]

            for cell, value in expected_headers:
                self.assertEqual(sheet[cell].value, value)

        workbook.close()
        os.remove(report_path)

    def test_report_data(self):
        """Проверка корректности данных в отчете"""
        report_path = generate_report(self.valid_data)
        workbook = load_workbook(report_path)
        sheet = workbook.active
        if sheet:
            expected_data = [
                ["Филиал 1", "Сотрудник 1", 5000000, 650000, 650000, 0],
                ["Филиал 2", "Сотрудник 2", 6000000, 800000, 800000, 0],
            ]

            for i, row in enumerate(
                sheet.iter_rows(min_row=3, max_row=4, values_only=True)
            ):
                self.assertEqual(list(row), expected_data[i])

        workbook.close()
        os.remove(report_path)

    def test_file_exists(self):
        """Проверка, что файл сохраняется корректно"""
        report_path = generate_report(self.valid_data)
        self.assertTrue(os.path.exists(report_path))
        os.remove(report_path)

    def test_empty_report(self):
        """Проверка создания отчета из пустого DataFrame"""
        empty_data = pd.DataFrame(
            {
                "Филиал": [],
                "Сотрудник": [],
                "Налоговая база": [],
                "Исчислено всего": [],
            }
        )
        with self.assertRaises(ValueError):
            generate_report(empty_data)


class ParseFileTests(TestCase):
    def create_excel_file(self, headers, data):
        workbook = Workbook()
        sheet = (
            workbook.active
            if workbook.active
            else workbook.create_sheet(title="Sheet1")
        )

        if not sheet:
            raise ValueError("Не удалось создать рабочий лист в Excel файле.")

        for i, header_row in enumerate(headers, start=1):
            for j, header in enumerate(header_row, start=1):
                sheet.cell(row=i, column=j).value = header

        for row_data in data:
            sheet.append(row_data)

        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        workbook.save(temp_file.name)
        temp_file.close()
        return temp_file.name

    def test_parse_valid_file(self):
        """Проверка, что файл с корректной структурой считывается правильно."""
        headers = [
            ["Филиал", "Сотрудник", "Доход", "Вычеты", "Налоговая база", "Налог"],
            [
                "",
                "",
                "Начислено",
                "Вычеты всего",
                "",
                "Исчислено всего",
                "Удержано всего",
            ],
        ]
        data = [
            ["Филиал 1", "Сотрудник 1", 1000000, 50000, 950000, 123500],
            ["Филиал 2", "Сотрудник 2", 2000000, 100000, 1900000, 247000],
        ]
        file_path = self.create_excel_file(headers, data)

        parsed_data = parse_file(file_path)

        self.assertEqual(len(parsed_data), 2)
        self.assertEqual(
            parsed_data.columns.tolist(),
            ["Филиал", "Сотрудник", "Налоговая база", "Исчислено всего"],
        )
        os.remove(file_path)

    def test_parse_file_with_missing_columns(self):
        """Проверка, что файл с недостаточным количеством столбцов вызывает ошибку."""
        headers = [
            ["Филиал", "Сотрудник", "Доход"],
            ["", "", "Начислено"],
        ]
        data = [["Филиал 1", "Сотрудник 1", 1000000]]
        file_path = self.create_excel_file(headers, data)

        with self.assertRaises(ValueError) as context:
            parse_file(file_path)
        self.assertEqual(
            str(context.exception), "Файл должен содержать минимум 6 столбцов."
        )
        os.remove(file_path)

    def test_parse_file_with_invalid_structure(self):
        """Проверка, что файл с некорректной структурой вызывает ошибку."""
        headers = [
            [
                "Некорректный",
                "Формат",
                "Заголовков",
                "Что-то ещё",
                "Еще что-то",
                "И ещё",
            ],
            ["", "", "", "", "", "", ""],
        ]
        data = [["Филиал 1", "Сотрудник 1", 1000000, 50000, 950000, 123500]]
        file_path = self.create_excel_file(headers, data)

        with self.assertRaises(ValueError) as context:
            parse_file(file_path)
        self.assertEqual(
            str(context.exception),
            "Файл имеет некорректную структуру. Проверьте исходные данные.",
        )
        os.remove(file_path)

    def test_parse_file_with_missing_required_columns(self):
        """Проверка файла, где отсутствуют обязательные данные."""
        headers = [
            ["Филиал", "Сотрудник", "Доход", "Вычеты", "Налоговая база", "Налог"],
            [
                "",
                "",
                "Начислено",
                "Вычеты всего",
                "",
                "Исчислено всего",
                "Удержано всего",
            ],
        ]
        data = [["Филиал 1", "Сотрудник 1", None, None, None, None]]
        file_path = self.create_excel_file(headers, data)

        with self.assertRaises(ValueError) as context:
            parse_file(file_path)
        self.assertEqual(
            str(context.exception),
            "Файл не содержит валидных данных после удаления пустых строк.",
        )
        os.remove(file_path)

    def test_parse_file_with_empty_data(self):
        """Проверка, что пустой файл вызывает ошибку."""
        headers = [
            ["Филиал", "Сотрудник", "Доход", "Вычеты", "Налоговая база", "Налог"],
            [
                "",
                "",
                "Начислено",
                "Вычеты всего",
                "",
                "Исчислено всего",
                "Удержано всего",
            ],
        ]
        data = []
        file_path = self.create_excel_file(headers, data)

        with self.assertRaises(ValueError) as context:
            parse_file(file_path)
        self.assertEqual(
            str(context.exception), "Файл пуст. Загрузка данных невозможна."
        )
        os.remove(file_path)
