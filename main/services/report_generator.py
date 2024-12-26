import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

logger = logging.getLogger(__name__)


def calculate_tax(base: float) -> float:
    logger.debug(f"Вычисление налога для базы: {base}")
    if base <= 5000000:
        tax = base * 0.13
    else:
        tax = 5000000 * 0.13 + (base - 5000000) * 0.15
    logger.debug(f"Рассчитанный налог: {tax}")
    return tax


def generate_report(data: pd.DataFrame) -> str:
    logger.info("Начало генерации отчета")
    try:
        data["Исчислено всего по формуле"] = data["Налоговая база"].apply(calculate_tax)
        logger.info("Добавлены расчеты налогов в DataFrame")

        data["Отклонения"] = (
            data["Исчислено всего"] - data["Исчислено всего по формуле"]
        )
        data = data.sort_values(by="Отклонения", ascending=False)
        logger.info("Данные отсортированы по отклонениям")

        wb = Workbook()
        ws = wb.active if wb.active else wb.create_sheet()

        # Оформление заголовков
        header_fill = PatternFill(
            start_color="CBE4E5", end_color="CBE4E5", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовки
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:B2")
        ws.merge_cells("C1:C2")
        ws.merge_cells("D1:E1")
        ws.merge_cells("F1:F2")

        ws["A1"].value = "Филиал"
        ws["B1"].value = "Сотрудник"
        ws["C1"].value = "Налоговая база"
        ws["D1"].value = "Налог"
        ws["D2"].value = "Исчислено всего"
        ws["E2"].value = "Исчислено всего по формуле"
        ws["F1"].value = "Отклонения"

        for col in range(1, 7):
            for row in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # Данные
        for _, row in data.iterrows():
            ws.append(row.tolist())

        # Форматирование отклонений
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=6, max_col=6):
            for cell in row:
                cell.border = border
                if cell.value == 0:
                    cell.fill = PatternFill(start_color="00FF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", fill_type="solid")

        # Автоширина столбцов
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Сохранение файла
        file_path = "report.xlsx"
        wb.save(file_path)
        logger.info(f"Отчет успешно сохранен: {file_path}")
        return file_path
    except Exception:
        logger.error("Ошибка генерации отчета", exc_info=True)
        raise
